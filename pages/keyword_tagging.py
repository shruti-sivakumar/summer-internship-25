import streamlit as st
import pandas as pd
from io import BytesIO
from keyword_tagger import load_model, process_excel_streaming
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("🏷️ Keyword Tagging for Bid Titles")
st.divider()

with st.expander("❗️ Instructions", expanded=False):
    st.markdown("""
    ### 🔹 **What This Tool Does**

    This tool uses a fine-tuned DistilRoBERTa model to predict relevant keywords for each bid title. It inserts a new column `AI Keywords` between `Data Miner Keywords` and `QC Keywords`.

    ### 📁 **Input Format**

    Upload an Excel (`.xlsx`) file with the following required columns:

    | Column Name           | Description                         |
    |------------------------|-------------------------------------|
    | `Bid Title`            | The bid title text to be tagged     |
    | `Data Miner Keywords`  | Existing manually assigned keywords |
    | `QC Keywords`          | Keywords after QC                   |

    ---

    ### ⚙️ **Steps to Use the Tool**

    1. **Upload Your Excel File**  
       Ensure the file includes headers at row 3 and data starts at row 5.

    2. **Processing**  
       The tool will:
       - Tag each bid title
       - Insert the `AI Keywords` column
       - Highlight confidence with colors

    3. **Download Output**  
       Preview and download the tagged Excel file.

    ---

    ### 🎨 **Confidence-Based Coloring**

    - 🟩 Green: High confidence (≥ 0.60)  
    - 🟨 Yellow: Medium confidence (0.40 – 0.59)  
    - 🟥 Red: Fallback / low confidence (< 0.40)
    """)

st.subheader("Upload Your Excel File")
uploaded_file = st.file_uploader("Format Given In Instructions", type=["xlsx"])

if uploaded_file:
    st.success("File uploaded. Click 'Upload' to start keyword tagging.")

    if st.button("Upload"):
        df = pd.read_excel(uploaded_file, header=2, skiprows=[3, 4])

        required_cols = {"Bid Title", "Data Miner Keywords", "QC Keywords"}
        if not required_cols.issubset(df.columns):
            st.error("Required columns missing. Please check your file.")
        else:
            progress = st.progress(0, text="Starting keyword tagging...")
            status = st.empty()

            model, tokenizer, keywords_list, mlb, device = load_model()

            output_df = df.copy()

            # Insert "AI Keywords" column between "Data Miner Keywords" and "QC Keywords"
            insert_index = output_df.columns.get_loc("Data Miner Keywords") + 1
            output_df.insert(insert_index, "AI Keywords", "")

            fallback_flags = []
            probabilities = []
            keywords_col = []

            for i, total, result in process_excel_streaming(output_df, model, tokenizer, keywords_list, device):
                keywords_col.append(result["keywords"])
                fallback_flags.append(result["fallback"])
                probabilities.append(result["max_prob"])

                percent = int((i / total) * 100)
                progress.progress(percent, f"Tagging row {i} of {total}...")

            # Apply final column after loop
            output_df["AI Keywords"] = keywords_col

            # Fix: convert all columns to string to avoid Arrow errors
            for col in output_df.columns:
                output_df[col] = output_df[col].astype(str)

            progress.progress(100, "Tagging complete!")
            status.success("All rows tagged.")

            st.dataframe(output_df, use_container_width=True)

            # Save Excel with coloring
            output_buffer = BytesIO()
            output_df.to_excel(output_buffer, index=False)
            output_buffer.seek(0)
            wb = load_workbook(output_buffer)
            ws = wb.active

            # Coloring
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Find column index for "AI Keywords"
            ai_col = None
            for cell in ws[1]:
                if cell.value == "AI Keywords":
                    ai_col = cell.col_idx
                    break

            if ai_col:
                for i, prob in enumerate(probabilities, start=2):
                    if prob >= 0.6:
                        ws.cell(row=i, column=ai_col).fill = green
                    elif prob >= 0.4:
                        ws.cell(row=i, column=ai_col).fill = yellow
                    else:
                        ws.cell(row=i, column=ai_col).fill = red

            final_buffer = BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="Download Tagged Excel",
                data=final_buffer,
                file_name="keyword_tagged_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )